#    Copyright 2022-2023 Daniel Balsom
#    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#    Permission is hereby granted, free of charge, to any person obtaining a
#    copy of this software and associated documentation files (the “Software”),
#    to deal in the Software without restriction, including without limitation
#    the rights to use, copy, modify, merge, publish, distribute, sublicense,
#    and/or sell copies of the Software, and to permit persons to whom the
#    Software is furnished to do so, subject to the following conditions:
#
#    The above copyright notice and this permission notice shall be included in
#    all copies or substantial portions of the Software.
#
#    THE SOFTWARE IS PROVIDED “AS IS”, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER   
#    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
#    FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
#    DEALINGS IN THE SOFTWARE.
#    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

#   excelify.py
#
#   Convert the specified input csv cycle log to Excel format using openpyxl.
#   Excel features such as cell borders and fill are used to make the output
#   pretty.
#
#   Command Line Arguments:
#   input_csv output_csv
import csv
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from collections import deque


PASTEL_PINK = 'FFD1DC'      # Pastel Pink
PASTEL_ORANGE = 'FFC3A0'    # Pastel Orange
PASTEL_YELLOW = 'FFF5A2'    # Pastel Yellow
PASTEL_GREEN = 'D3FFA3'     # Pastel Green
PASTEL_MINT = 'A2F9E6'      # Pastel Mint
PASTEL_SKY_BLUE = 'A2D5F9'  # Pastel Sky Blue
PASTEL_PURPLE = 'CCA2F9'    # Pastel Purple
PASTEL_VIOLET = 'FCA2F9'    # Pastel Violet
LIGHT_PINK = 'FFB6C1'       # Light Pink
LIGHT_BLUE = 'ADD8E6'       # Light Blue
PALE_GREEN = '98FB98'       # Pale Green
LIGHT_GOLDEN = 'FFD700'     # Light Golden
WHEAT = 'F5DEB3'            # Wheat
LIGHT_CYAN = 'E0FFFF'       # Light Cyan
THISTLE = 'D8BFD8'          # Thistle
SANDY_BROWN = 'F4A460'      # Sandy Brown
LIGHT_GRAY = 'D3D3D3'  # Light Gray
NORMAL_BLUE = '0000FF'

FILL_COLORS = ["ADD8E6", "FFB6C1", "FFDAB9", "E0FFFF", "FFF0F5", "D8BFD8", "FAEBD7"]
COLOR_MAP = {
    'SEG': {
        #'CS': LIGHT_BLUE,
        'DS': PASTEL_GREEN,
        'SS': PASTEL_PINK,
        'ES': PASTEL_YELLOW
    },
    'T': {
        'TI': LIGHT_GRAY,
        'Tw': LIGHT_BLUE
    },
    'QOP': {
        'F': PASTEL_MINT,
        'S': PASTEL_YELLOW,
        'E': PASTEL_PINK,
    }
}

def apply_fill(wb):
    """
    Iterates through the worksheet and applies fill color based on column header and cell value.

    :param ws: The worksheet object.
    :param color_map: The color mapping.
    """

    print("\nApplying cell fills...")

    ws = wb.active
    headers = {}
    for col_num, cell in enumerate(ws[1], 1):  # assuming headers are in row 1
        headers[col_num] = cell.value

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # starting from row 2
        for col_num, cell_value in enumerate(row, 1):
            col_header = headers[col_num]
            if col_header in COLOR_MAP and cell_value in COLOR_MAP[col_header]:
                color_hex = COLOR_MAP[col_header][cell_value]
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                ws.cell(row=row_num, column=col_num).fill = fill

def set_clk_borders(ws, curr, prev, nxt, row, col_num, fill_color):

    fill_style = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    borders = {"left": None, "right": None, "top": None, "bottom": None}
    do_fill = None

    if curr != prev:
        borders["top"] = Side(style='thin')
    if curr == 0:
        borders["left"] = Side(style='thin')
    if curr == 1:
        borders["right"] = Side(style='thin')
        do_fill = fill_style
    if curr != nxt:
        borders["bottom"] = Side(style='thin')

    ws.cell(row=row, column=col_num + 2).border = Border(
        left=borders["left"],
        right=borders["right"],
        top=borders["top"],
        bottom=borders["bottom"]
    )

    if do_fill:
        ws.cell(row=row, column=col_num + 2).fill = do_fill

def adjust_column_width(ws, col_num, skinny=True):

    col_letter = get_column_letter(col_num)
    default_height = 15.0
    row_height_points = ws.row_dimensions[1].height or default_height

    print(f"Adjusting width for column {col_letter}")

    if skinny:
        ws.column_dimensions[col_letter].width = row_height_points / 10
    else:
        ws.column_dimensions[col_letter].width = row_height_points / 5


def read_csv(csv_filename):

    with open(csv_filename, mode='r') as csv_file:
        reader = csv.reader(csv_file)
        headers = next(reader)
        data = [row for row in reader]

    return headers, data

def csv_to_excel(csv_filename):
    
    wb = Workbook()
    ws = wb.active
    instructions = deque()

    with open(csv_filename, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)

        print("Creating excel worksheet...")

        for idx, header in enumerate(csv_reader.fieldnames, start=1):
            ws.cell(row=1, column=idx, value=header)

        for i, row in enumerate(csv_reader):
            if i % 1000 == 0:
                sys.stdout.write(f'\rProcessed row: {i}')
                sys.stdout.flush()
            ws.append(list(row.values()))

            if row['DISASM']:
                #print(f"Found instruction at: {i}")
                instructions.append(i + 3)

    return wb, instructions

def find_disasm_rows(headers, data):

    disasm_rows = []
    try:
        disasm_index = headers.index('DISASM')
    except ValueError:
        raise ValueError("No 'DISASM' column found in the CSV file.")
    
    for i, row in enumerate(data):
        if row[disasm_index]:
            disasm_rows.append(i)

    return disasm_rows

def set_border(cell, row_num, disasm_rows):

    if (row_num - 2) in disasm_rows:
        thick_border_side = Side(style='thick')
        cell.border = Border(top=thick_border_side)

def set_instr_border(cell):
    thick_border = Side(style='thick')

    left = cell.border.left
    right = cell.border.right
    bottom = cell.border.bottom

    cell.border = Border(left=left, right=right, top=thick_border, bottom=bottom)

def set_fill(cell, value):

    blue_fill = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    if value == 'Tw':
        cell.fill = blue_fill
    elif value == 'TI':
        cell.fill = grey_fill

def add_instruction_borders(wb, instructions):

    print("\nAdding instruction borders...")

    ws = wb.active

    have_disasm = False

    for cell in ws[1]:  # assuming headers are in the first row
        if cell.value == 'DISASM':
            disasm_index = cell.column  # get the column number
            have_disasm = True
            break

    if not have_disasm:
        return
    
    next_instr_idx = instructions.popleft()

    # Mark instructions with horizontal borders.
    for i, row in enumerate(ws.iter_rows(), start=2): 

        if i % 1000 == 0:
            sys.stdout.write(f'\rProcessed row: {i}')
            sys.stdout.flush()
        
        if i == next_instr_idx:
            #print(f"\nMarking instruction: {i}")
            for col_num, cell in enumerate(row, start=1):
                set_instr_border(cell)
            try:
                next_instr_idx = instructions.popleft()
            except:
                return
                
                
def draw_clocks(wb, columns):

    color_index = 0

    print("\nDrawing clocks...")
    ws = wb.active
    
    # Iterate over column headers to find clk columns
    col_num = 1

    # Add columns to represent clock art
    for cell in ws[1]:
        cell_value = cell.value
        # If the header value matches our list
        if str(cell_value).upper() in columns:
            # Insert two new columns to the right of the clock data column
            ws.insert_cols(col_num + 1, 2)
            
            ws.cell(row=1, column=col_num + 1, value=str(cell_value).upper() + "1")
            ws.cell(row=1, column=col_num + 2, value=str(cell_value).upper() + "2")

            adjust_column_width(ws, col_num + 1)
            adjust_column_width(ws, col_num + 2, False)
            
            # Move to the next column and increment color index
            col_num += 3
        else:
            col_num += 1

    # Reset column number and actually draw them
    col_num = 0

    for col_num, cell in enumerate(ws[1]):

        cell_value = cell.value
        #print(f"Evaluating column: {str(cell_value).upper()}")

        # If the header value matches our list
        if str(cell_value).upper() in columns:

            print(f"Drawing clock column: {cell_value}")

            start_row = 2  # Assuming headers are present
            clk_data = [
                value
                for value in ws.iter_cols(
                    min_col=col_num + 1,
                    max_col=col_num + 1,
                    min_row=start_row,
                    values_only=True,
                )
            ][0]
            
            # Insert two new columns to the right of the clock data column
            #ws.insert_cols(col_num + 1, 2)

            # Set width for the first inserted column
            
            #ws.column_dimensions[chr(65 + col_num)].width = 1.0  # Static small width
            #adjust_column_width(ws, col_num)

            for index, value in enumerate(clk_data, start_row):
                if not value:
                    print(f"No value at index: {index}")
                    break

                value = int(value)
                prev_value = (
                    int(clk_data[index - start_row - 1]) if index > start_row else value
                )
                next_value = (
                    int(clk_data[index - start_row + 1])
                    if index - start_row < len(clk_data) - 1
                    else value
                )
                
                # Only set fill color if the value is 0 or 1
                if value in [0, 1]:

                    ws.cell(row=index, column=col_num + 2).fill = PatternFill(
                        start_color=FILL_COLORS[color_index],
                        end_color=FILL_COLORS[color_index],
                        fill_type="solid",
                    )
                    set_clk_borders(
                        ws,
                        value,
                        prev_value,
                        next_value,
                        index,
                        col_num + 1,
                        FILL_COLORS[color_index],
                    )

            # Adjust column width after drawing the clock
            #adjust_column_width(ws, col_num + 1)
            
            # Move to the next column and increment color index
            col_num += 3
            color_index = (color_index + 1) % len(FILL_COLORS)

def convert_columns_to_text(wb, col_headers):
    """
    Converts specified columns to text and removes any "`" character from the values.

    :param ws: The worksheet object.
    :param col_headers: List of column header names to be converted.
    """

    ws = wb.active
    # First, identify the column numbers for the headers.
    header_col_nums = {}
    for col_num, cell in enumerate(ws[1], 1):  # assuming headers are in row 1
        if cell.value in col_headers:
            header_col_nums[cell.value] = col_num
    
    # Next, iterate over the cells in the specified columns.
    for col_header, col_num in header_col_nums.items():
        for row_num, row in enumerate(ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num), start=2):  # starting from row 2
            cell = ws.cell(row=row_num, column=col_num)
            # If cell contains a value, convert it to text and remove any "`" character.
            if cell.value is not None:
                cell.value = str(cell.value).replace("'", "")

def adjust_column_widths(wb, ignore_cols, padding=2):
    """
    Iterate through columns in the worksheet and set the column width based on the
    length of the longest string in each column.

    :param ws: The worksheet object.
    :param padding: An optional padding to add to the column width.
    """

    ws = wb.active

    min_width = 1

    skip_cols = []

    # Add clock columns
    for item in ignore_cols:
        skip_cols.append(item)
        skip_cols.append(item + '1')
        skip_cols.append(item + '2')

    for col_idx, col in enumerate(ws.columns, 1):  # starting enumeration from 1
        max_length = 0
        # If the header of the column is in skip_cols, skip the column
        
        if col[0].value in skip_cols:
            print(f"skipping resize of column: {col[0].value}")
            continue

        for cell in col[1:]:  # skip the header cell
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        if max_length > 0:  # Only adjust column width if a nonzero length string was found.
            adjusted_width = max_length + padding
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            print(f"resizing column: {col[0].value} to width: {adjusted_width}")
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = min_width

def remove_columns_by_header(wb, headers):
    """
    Removes columns from the worksheet based on the header names.

    :param workbook: The loaded workbook object.
    :param headers: A list of header names to be removed.
    """
    ws = wb.active  # assuming we are working on the active sheet
    header_row = [cell.value for cell in ws[1]]  # assuming headers are in the first row
    
    # Get column indices for headers to be removed
    col_indices = [header_row.index(header) + 1 for header in headers if header in header_row]

    # Sort indices in descending order and remove columns
    for col_idx in sorted(col_indices, reverse=True):
        ws.delete_cols(col_idx)

def keep_only_columns(wb, headers_to_keep):
    """
    Keeps only the columns from the worksheet based on the header names provided and removes all other columns.

    :param workbook: The loaded workbook object.
    :param headers_to_keep: A list of header names to be kept.
    """
    ws = wb.active  # assuming we are working on the active sheet
    header_row = [cell.value for cell in ws[1]]  # assuming headers are in the first row

    # Find headers to remove: all headers that are not in headers_to_keep
    headers_to_remove = [header for header in header_row if header not in headers_to_keep]

    # Call the remove_columns_by_header function to remove these headers
    remove_columns_by_header(wb, headers_to_remove)

def print_cols(wb):

    ws = wb.active
    for cell in ws[1]:
        print(f"column name: {cell.value}")

def add_instructions_sheet(wb):

    print("Adding instruction index...")
    # Assuming the first sheet is the main sheet
    ws = wb.active

    # Create 'Instructions' sheet if not exist
    if 'Instructions' in wb.sheetnames:
        instructions_sheet = wb['Instructions']
    else:
        instructions_sheet = wb.create_sheet(title='Instructions')

    # Find the index of the "DISASM" column in the main sheet
    disasm_col = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'DISASM':
            disasm_col = col
            break
    
    if disasm_col is None:
        raise Exception("DISASM column not found in the main sheet")
    
    # Fixed-width font style and hyperlink font style
    fixed_font = Font(name='Courier New')
    hyperlink_font = Font(color='0000FF', underline='single')

    # Variables for tracking cycles
    last_row_num = 0
    row_num = 2  # Start from the second row
    first_cycle = True

    # Adding headers
    instructions_sheet['A1'] = 'DISASM'
    instructions_sheet['B1'] = 'CYCLES'  # Swapped column
    instructions_sheet['C1'] = 'LINK'    # Swapped column

    # Iterate through cells in "DISASM" column in the main sheet starting from row 2
    for index, cell in enumerate(disasm_col[1:], start=2):
        # If cell has a value, copy it to "Instructions" sheet and create a link
        if cell.value:
            instructions_sheet[f'A{row_num}'].value = cell.value
            instructions_sheet[f'A{row_num}'].font = fixed_font
            instructions_sheet[f'C{row_num}'] = f'=HYPERLINK("#\'{ws.title}\'!{cell.coordinate}", "Link to {cell.coordinate}")'
            instructions_sheet[f'C{row_num}'].font = hyperlink_font
            
            # Calculate the number of cycles (rows between instructions)
            if last_row_num != 1:
                cycles = index - last_row_num
                if first_cycle:
                    first_cycle = False
                else:
                    instructions_sheet[f'B{row_num - 1}'] = cycles
            last_row_num = index
            row_num += 1

def add_io_sheet(wb):

    print("Adding IO index...")

    try:
        # Try to open and load the ports.json file
        with open('ports.json', 'r') as file:
            ports = json.load(file)
    except FileNotFoundError:
        print("Error: The ports.json file could not be found.")
        ports = {}  # Set ports to an empty dictionary
    except json.JSONDecodeError:
        print("Error: Unable to decode JSON data in ports.json file.")
        ports = {}  # Set ports to an empty dictionary

    main_sheet = wb.active

    # Create 'IO' sheet if not exist
    if 'IO' in wb.sheetnames:
        io_sheet = wb['IO']
    else:
        io_sheet = wb.create_sheet(title='IO')

    # Adding headers
    io_sheet['A1'] = 'ADDR'
    io_sheet['B1'] = 'OP'
    io_sheet['C1'] = 'DATA'
    io_sheet['D1'] = 'DESC'  # Adding a 'DESC' column header

    # Initialize row_num for IO sheet
    row_num = 2

    # Resolve the column headers to column numbers
    busl_col, al_col, d_col = None, None, None
    for i, col in enumerate(main_sheet.iter_cols(1, main_sheet.max_column)):
        if col[0].value == 'BUSL':
            busl_col = i + 1
        elif col[0].value == 'AL':
            al_col = i + 1
        elif col[0].value == 'D':
            d_col = i + 1

    if not busl_col or not al_col or not d_col:
        raise Exception("Required columns not found in the main sheet")            

    # Iterate through cells in 'D' column in the main sheet starting from row 2
    for row in main_sheet.iter_rows(min_row=2, min_col=d_col, max_col=d_col, max_row=main_sheet.max_row):
        for cell in row:
            if cell.value:  # If cell in 'D' column has a value
                busl_value = main_sheet.cell(row=cell.row, column=busl_col).value  # BUSL value in the same row
                al_value = main_sheet.cell(row=cell.row, column=al_col).value  # AL value in the same row
                
                if busl_value == 'IOR' or busl_value == 'IOW':
                    io_sheet[f'A{row_num}'] = al_value
                    io_sheet[f'B{row_num}'] = 'R' if busl_value == 'IOR' else 'W'
                    io_sheet[f'C{row_num}'] = cell.value

                    # Construct the key and look up in the ports.json
                    addr = al_value[-4:]
                    op = io_sheet[f'B{row_num}'].value
                    key = f"{addr}{op.lower()}"
                    desc = ports.get(key, '')  # If not found, default to an empty string
                    io_sheet[f'D{row_num}'] = desc

                    row_num += 1

    # Set the auto_filter to include all data in the IO sheet.
    io_sheet.auto_filter.ref = io_sheet.dimensions

def main():

    clk_columns = ["READY", "CLK0", "INTR", "DR0", "VS", "HS"]
    text_columns = ["AL", "INSTF", "D", "QB", "Q0", "Q1", "Q2", "Q3", "ADDR"]
    keep_columns = [
        'N', 'ALE', 'AL', 'SEG', 'BUSL', 'READY', 'READY1', 'READY2', 'T', 'D', 'QOP', 'QB', 'INSTF', 'DISASM', 
        'QL', 'Q0', 'Q1', 'Q2', 'Q3',
        'CLK0', 'CLK01', 'CLK02', 'INTR', 'INTR1', 'INTR2', 'DR0', 'DR01', 'DR02', 'VS', 'VS1', 'VS2', 'HS', 'HS1', 'HS2',
        "FRAME", "R_X", "R_Y"
    ]
    
    if len(sys.argv) != 3:
        print("Usage: python excelify.py <input.csv> <output.xlsx>")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_xlsx = sys.argv[2]
    
    #headers, data = read_csv(input_csv)

    wb, instructions = csv_to_excel(input_csv)

    print(f"\nFound {len(instructions)} instructions...")
    
    keep_only_columns(wb, keep_columns)
    draw_clocks(wb, clk_columns)
    adjust_column_widths(wb, clk_columns, padding=4)
    add_instruction_borders(wb, instructions)
    
    convert_columns_to_text(wb, text_columns)
    apply_fill(wb)
    
    # Freeze the top row
    wb.active.freeze_panes = 'A2'

    add_instructions_sheet(wb)
    add_io_sheet(wb)

    print("\nSaving xlsx...")
    wb.save(output_xlsx)
    #disasm_rows = find_disasm_rows(headers, data)
    #write_to_xlsx(headers, data, output_xlsx, disasm_rows)
    #wb.save(output_xlsx)

if __name__ == "__main__":
    main()
